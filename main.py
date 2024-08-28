import telebot
from telebot import types
from openpyxl import load_workbook

Token = '7104291510:AAE5Neqcop9-fAo3LcFkb1AmA0y0NBPxu24'
bot = telebot.TeleBot(Token)
fn = 'tablich.xlsx'
wb = load_workbook(fn)
ws = wb['data']

# ws['A1'] = 'Qq'
# print(len(ws['A1'].value))
@bot.message_handler(content_types=['text'])
def main(message):  # имя
    if message.text == '/start':
        bot.send_message(message.from_user.id, f'Наш чат-бот в Telegram предназначен специально для молодых семей-студентов, которые желают получить психологическое консультирование по вопросам брачно-семейных отношений🫂'
        f'‍✍️Пройдите первичное анкетирование, чтобы мы могли лучше понять ваши потребности и предоставить вам персонализированную помощь. '
        f'👩‍⚕️Наша команда психологов готова поддержать вас на пути к здоровым и счастливым отношениям.')
        bot.send_message(message.from_user.id, f'Напишите ваше полное имя')
        bot.register_next_step_handler(message, old)
    else:
        bot.send_message(message.chat.id, 'Напишите /start, чтобы начать опрос')


def old(message):  # возвраст
    # поиск подходящего номера для exel
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    if ws['A2'].value == None:
        ws['A2'] = message.text
    else:
        s = 2
        while True:
            if ws['A' + str(s)].value == None:
                break
            s += 1
        ws['A' + str(s)] = message.text
        ws['A1'] = str(s)
    wb.save(fn)
    wb.close()
    #
    bot.send_message(message.chat.id, 'Сколько вам лет?')
    bot.register_next_step_handler(message, gender)


def gender(message):  # пол
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['B' + ws['A1'].value] = str(message.text)
    wb.save(fn)
    wb.close()
    #
    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton('Мужской')
    btn2 = types.KeyboardButton('Женский')
    markup.add(btn1, btn2)
    bot.send_message(message.chat.id, 'Ваш пол?', reply_markup=markup)
    bot.register_next_step_handler(message, family)


def family(message):  # семейное положение
    if message.text.lower() == 'мужской' or message.text.lower() == 'женский':
        # использование этого номера
        fn = 'tablich.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['C' + ws['A1'].value] = message.text
        wb.save(fn)
        wb.close()
        #
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Состою в браке')
        btn2 = types.KeyboardButton('Планирую вступить в брак')
        markup.add(btn1, btn2)
        bot.send_message(message.chat.id, 'Вы состоите или планируете вступить в брак?', reply_markup=markup)
        bot.register_next_step_handler(message, children)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Да')
        btn2 = types.KeyboardButton('Нет')
        bot.send_message(message.chat.id, 'Ваш пол?', reply_markup=markup)
        bot.register_next_step_handler(message, family)


def children(message):  # дети
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['D' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'состою в браке':
        markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, 'Как долго вы состоите в браке?', reply_markup=markup)
        bot.register_next_step_handler(message, family_old)
    elif message.text.lower() == 'планирую вступить в брак':
        # использование этого номера
        fn = 'tablich.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['E' + ws['A1'].value] = '-'
        wb.save(fn)
        wb.close()
        #
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Да')
        btn2 = types.KeyboardButton('Нет')
        markup.add(btn1, btn2)
        bot.send_message(message.chat.id, 'Есть ли у вас дети?', reply_markup=markup)
        bot.register_next_step_handler(message, education)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Состою в браке')
        btn2 = types.KeyboardButton('Планирую вступить в брак')
        markup.add(btn1, btn2)
        bot.send_message(message.chat.id, 'Вы состоите или планируете вступить в брак?', reply_markup=markup)
        bot.register_next_step_handler(message, children)


def family_old(message):  # срок отношений
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['E' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton('Да')
    btn2 = types.KeyboardButton('Нет')
    markup.add(btn1, btn2)
    bot.send_message(message.chat.id, 'Есть ли у вас дети?', reply_markup=markup)
    bot.register_next_step_handler(message, education)


def education(message):  # образование
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['F' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'да':
        markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, 'Сколько детей у вас? И сколько им лет?', reply_markup=markup)
        bot.register_next_step_handler(message, children_old)
    elif message.text.lower() == 'нет':
        # использование этого номера
        fn = 'tablich.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['G' + ws['A1'].value] = '-'
        wb.save(fn)
        wb.close()
        ##111
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Среднее')
        btn2 = types.KeyboardButton('Неполное высшее')
        btn3 = types.KeyboardButton('Высшее')
        markup.add(btn1, btn2, btn3)
        bot.send_message(message.chat.id, 'Какой у вас уровень образования?', reply_markup=markup)
        bot.register_next_step_handler(message, educational_institution)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Да')
        btn2 = types.KeyboardButton('Нет')
        markup.add(btn1, btn2)
        bot.send_message(message.chat.id, 'Есть ли у вас дети?', reply_markup=markup)
        bot.register_next_step_handler(message, education)


def children_old(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['G' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton('Среднее')
    btn2 = types.KeyboardButton('Неполное высшее')
    btn3 = types.KeyboardButton('Высшее')
    markup.add(btn1, btn2, btn3)
    bot.send_message(message.chat.id, 'Какой у вас уровень образования?', reply_markup=markup)
    bot.register_next_step_handler(message, educational_institution)


def educational_institution(message):  # учебное заведение
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['H' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text == 'Среднее' or message.text == 'Неполное высшее' or message.text == 'Высшее':
        markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, 'В каком учебном заведении вы обучаетесь или обучались?', reply_markup=markup)
        bot.register_next_step_handler(message, department_and_course)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Среднее')
        btn2 = types.KeyboardButton('Неполное высшее')
        btn3 = types.KeyboardButton('Высшее')
        markup.add(btn1, btn2, btn3)
        bot.send_message(message.chat.id, 'Какой у вас уровень образования?', reply_markup=markup)
        bot.register_next_step_handler(message, educational_institution)

def department_and_course(message):  # факультет и курс
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['I' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    markup = types.ReplyKeyboardRemove()
    bot.send_message(message.chat.id, ' fльтете и курсе вы учитесь или учились?', reply_markup=markup)
    bot.register_next_step_handler(message, work)


def work(message):  # работа
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['J' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton('Да')
    btn2 = types.KeyboardButton('Нет')
    markup.add(btn1, btn2)
    bot.send_message(message.chat.id, 'Работаете ли вы в настоящее время?', reply_markup=markup)
    bot.register_next_step_handler(message, Combining_study_and_work)


def Combining_study_and_work(message):  # совмещение работы и учебы
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['K' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'да':
        markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, 'Где вы работаете? И на какой должности?', reply_markup=markup)
        bot.register_next_step_handler(message, Job_and_position)
    elif message.text.lower() == 'нет':
        # использование этого номера
        fn = 'tablich.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['L' + ws['A1'].value] = '-'
        wb.save(fn)
        wb.close()
        #
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень трудно')
        btn2 = types.KeyboardButton('Трудно')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Легко')
        btn5 = types.KeyboardButton('Очень легко')
        btn6 = types.KeyboardButton('Не приходится совмещать')
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6)
        bot.send_message(message.chat.id, 'Как вы справляетесь с совмещением учебы и работы?', reply_markup=markup)
        bot.register_next_step_handler(message, stress)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Да')
        btn2 = types.KeyboardButton('Нет')
        markup.add(btn1, btn2)
        bot.send_message(message.chat.id, 'Работаете ли вы в настоящее время?', reply_markup=markup)
        bot.register_next_step_handler(message, Combining_study_and_work)


def Job_and_position(message):  # работа и должность
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['L' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton('Очень трудно')
    btn2 = types.KeyboardButton('Трудно')
    btn3 = types.KeyboardButton('Нейтрально')
    btn4 = types.KeyboardButton('Легко')
    btn5 = types.KeyboardButton('Очень легко')
    btn6 = types.KeyboardButton('Не приходится совмещать')
    markup.add(btn1, btn2, btn3, btn4, btn5, btn6)
    bot.send_message(message.chat.id, 'Как вы справляетесь с совмещением учебы и работы?', reply_markup=markup)
    bot.register_next_step_handler(message, stress)


def stress(message):  # стресс
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['M' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'очень трудно' or message.text.lower() == 'трудно' or message.text.lower() == 'нейтрально' or message.text.lower() == 'легко' or message.text.lower() == 'очень легко' or message.text.lower() == 'не приходится совмещать':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как часто вы испытываете стресс?', reply_markup=markup)
        bot.register_next_step_handler(message, Anxiety)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень трудно')
        btn2 = types.KeyboardButton('Трудно')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Легко')
        btn5 = types.KeyboardButton('Очень легко')
        btn6 = types.KeyboardButton('Не приходится совмещать')
        markup.add(btn1, btn2, btn3, btn4, btn5, btn6)
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как вы справляетесь с совмещением учебы и работы?', reply_markup=markup)
        bot.register_next_step_handler(message, stress)


def Anxiety(message):  # тревожность
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['N' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'никогда' or message.text.lower() == 'редко' or message.text.lower() == 'иногда' or message.text.lower() == 'часто' or message.text.lower() == 'постоянно':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как часто вы испытываете тревогу?', reply_markup=markup)
        bot.register_next_step_handler(message, Depression)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как часто вы испытываете стресс?', reply_markup=markup)
        bot.register_next_step_handler(message, Anxiety)


def Depression(message):  # депрессия
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['O' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'никогда' or message.text.lower() == 'редко' or message.text.lower() == 'иногда' or message.text.lower() == 'часто' or message.text.lower() == 'постоянно':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Замечали ли вы у себя симптомы депрессии?', reply_markup=markup)
        bot.register_next_step_handler(message, Self_esteem)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как часто вы испытываете тревогу?', reply_markup=markup)
        bot.register_next_step_handler(message, Depression)


def Self_esteem(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['P' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'никогда' or message.text.lower() == 'редко' or message.text.lower() == 'иногда' or message.text.lower() == 'часто' or message.text.lower() == 'постоянно':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень низкая')
        btn2 = types.KeyboardButton('Низкая')
        btn3 = types.KeyboardButton('Средняя')
        btn4 = types.KeyboardButton('Высокая')
        btn5 = types.KeyboardButton('Очень высокая')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как вы оцениваете свою самооценку?', reply_markup=markup)
        bot.register_next_step_handler(message, Social_support)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Замечали ли вы у себя симптомы депрессии?', reply_markup=markup)
        bot.register_next_step_handler(message, Self_esteem)


def Social_support(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['Q' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'очень низкая' or message.text.lower() == 'низкая' or message.text.lower() == 'средняя' or message.text.lower() == 'высокая' or message.text.lower() == 'очень высокая':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Чувствуете ли вы поддержку со стороны семьи и друзей?', reply_markup=markup)
        bot.register_next_step_handler(message, Satisfaction_with_life)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень низкая')
        btn2 = types.KeyboardButton('Низкая')
        btn3 = types.KeyboardButton('Средняя')
        btn4 = types.KeyboardButton('Высокая')
        btn5 = types.KeyboardButton('Очень высокая')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как вы оцениваете свою самооценку?', reply_markup=markup)
        bot.register_next_step_handler(message, Social_support)


def Satisfaction_with_life(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['R' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'никогда' or message.text.lower() == 'редко' or message.text.lower() == 'иногда' or message.text.lower() == 'часто' or message.text.lower() == 'постоянно':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень недоволен')
        btn2 = types.KeyboardButton('Недоволен')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Доволен')
        btn5 = types.KeyboardButton('Очень доволен')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько вы довольны своей жизнью в целом? ', reply_markup=markup)
        bot.register_next_step_handler(message, Conflicts)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Чувствуете ли вы поддержку со стороны семьи и друзей?', reply_markup=markup)
        bot.register_next_step_handler(message, Satisfaction_with_life)


def Conflicts(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['S' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'очень недоволен' or message.text.lower() == 'недоволен' or message.text.lower() == 'нейтрально' or message.text.lower() == 'доволен' or message.text.lower() == 'очень доволен':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как часто у вас возникают конфликты с партнером?', reply_markup=markup)
        bot.register_next_step_handler(message, Conflict_resolution)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень недоволен')
        btn2 = types.KeyboardButton('Недоволен')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Доволен')
        btn5 = types.KeyboardButton('Очень доволен')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько вы довольны своей жизнью в целом? ', reply_markup=markup)
        bot.register_next_step_handler(message, Conflicts)


def Conflict_resolution(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['T' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'никогда' or message.text.lower() == 'редко' or message.text.lower() == 'иногда' or message.text.lower() == 'часто' or message.text.lower() == 'постоянно':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень трудно')
        btn2 = types.KeyboardButton('Трудно')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Легко')
        btn5 = types.KeyboardButton('Очень легко')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько легко вам удается решать конфликты?', reply_markup=markup)
        bot.register_next_step_handler(message, Trust)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как часто у вас возникают конфликты с партнером?', reply_markup=markup)
        bot.register_next_step_handler(message, Conflict_resolution)


def Trust(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['U' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'очень трудно' or message.text.lower() == 'трудно' or message.text.lower() == 'нейтрально' or message.text.lower() == 'легко' or message.text.lower() == 'очень легко':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Совсем не доверяю')
        btn2 = types.KeyboardButton('Не доверяю')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Доверяю')
        btn5 = types.KeyboardButton('Полностью доверяю')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько вы доверяете своему партнеру?', reply_markup=markup)
        bot.register_next_step_handler(message, Emotional_intimacy)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень трудно')
        btn2 = types.KeyboardButton('Трудно')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Легко')
        btn5 = types.KeyboardButton('Очень легко')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько легко вам удается решать конфликты?', reply_markup=markup)
        bot.register_next_step_handler(message, Trust)


def Emotional_intimacy(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['V' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'совсем не доверяю' or message.text.lower() == 'не доверяю' or message.text.lower() == 'нейтрально' or message.text.lower() == 'доверяю' or message.text.lower() == 'полностью доверяю':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень низкая')
        btn2 = types.KeyboardButton('Низкая')
        btn3 = types.KeyboardButton('Средняя')
        btn4 = types.KeyboardButton('Высокая')
        btn5 = types.KeyboardButton('Очень высокая')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как вы оцениваете вашу эмоциональную близость с партнером?',
                         reply_markup=markup)
        bot.register_next_step_handler(message, Time_together)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Совсем не доверяю')
        btn2 = types.KeyboardButton('Не доверяю')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Доверяю')
        btn5 = types.KeyboardButton('Полностью доверяю')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько вы доверяете своему партнеру?', reply_markup=markup)
        bot.register_next_step_handler(message, Emotional_intimacy)


def Time_together(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['W' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'очень низкая' or message.text.lower() == 'низкая' or message.text.lower() == 'средняя' or message.text.lower() == 'высокая' or message.text.lower() == 'очень высокая':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Меньше 1 часа в день')
        btn2 = types.KeyboardButton('1-2 часа в день')
        btn3 = types.KeyboardButton('3-4 часа в день')
        btn4 = types.KeyboardButton('более 4 часов в день')
        markup.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, 'Сколько времени вы проводите вместе с партнером?', reply_markup=markup)
        bot.register_next_step_handler(message, Roles_in_the_family)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень низкая')
        btn2 = types.KeyboardButton('Низкая')
        btn3 = types.KeyboardButton('Средняя')
        btn4 = types.KeyboardButton('Высокая')
        btn5 = types.KeyboardButton('Очень высокая')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как вы оцениваете вашу эмоциональную близость с партнером?',
                         reply_markup=markup)
        bot.register_next_step_handler(message, Time_together)


def Roles_in_the_family(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['X' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'меньше 1 часа в день' or message.text.lower() == '1-2 часа в день' or message.text.lower() == '3-4 часа в день' or message.text.lower() == 'более 4 часов в день':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Равномерно')
        btn2 = types.KeyboardButton('Неравномерно, большинство обязанностей у меня')
        btn3 = types.KeyboardButton('Неравномерно, большинство обязанностей у партнера')
        markup.add(btn1, btn2, btn3)
        bot.send_message(message.chat.id, 'Как вы распределяете роли и обязанности в семье?', reply_markup=markup)
        if ws['D' + ws['A1'].value].value == 'Планирую вступить в брак':
            bot.register_next_step_handler(message, Motivation_for_marriage)
        else:
            bot.register_next_step_handler(message, Financial_stability)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Меньше 1 часа в день')
        btn2 = types.KeyboardButton('1-2 часа в день')
        btn3 = types.KeyboardButton('3-4 часа в день')
        btn4 = types.KeyboardButton('более 4 часов в день')
        markup.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, 'Сколько времени вы проводите вместе с партнером?', reply_markup=markup)
        bot.register_next_step_handler(message, Roles_in_the_family)


# Раздел 5: Готовность к браку (для тех, кто планирует вступить в брак)
def Motivation_for_marriage(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['Y' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'равномерно' or message.text.lower() == 'неравномерно, большинство обязанностей у меня' or message.text.lower() == 'неравномерно, большинство обязанностей у партнера':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Любовь')
        btn2 = types.KeyboardButton('Социальные нормы')
        btn3 = types.KeyboardButton('Материальная поддержка')
        btn4 = types.KeyboardButton('Другие причины')
        markup.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, 'Почему вы решили вступить в брак? ', reply_markup=markup)
        bot.register_next_step_handler(message, Expectations_of_marriage)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Равномерно')
        btn2 = types.KeyboardButton('Неравномерно, большинство обязанностей у меня')
        btn3 = types.KeyboardButton('Неравномерно, большинство обязанностей у партнера')
        markup.add(btn1, btn2, btn3)
        bot.send_message(message.chat.id, 'Как вы распределяете роли и обязанности в семье?', reply_markup=markup)
        bot.register_next_step_handler(message, Motivation_for_marriage)


def Expectations_of_marriage(message):
    if message.text.lower() == 'любовь' or message.text.lower() == 'социальные нормы' or message.text.lower() == 'материальная поддержка':
        # использование этого номера
        fn = 'tablich.xlsx'
        wb = load_workbook(fn)
        ws = wb['data']
        ws['Z' + ws['A1'].value] = message.text
        wb.save(fn)
        wb.close()
        #
        markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, 'Каковы ваши ожидания от брака?', reply_markup=markup)
        bot.register_next_step_handler(message, Partner_support)
    elif message.text.lower() == 'другие причины':
        markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, 'Какие причины?', reply_markup=markup)
        bot.register_next_step_handler(message, Expectations_of_marriage1)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Любовь')
        btn2 = types.KeyboardButton('Социальные нормы')
        btn3 = types.KeyboardButton('Материальная поддержка')
        btn4 = types.KeyboardButton('Другие причины')
        markup.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, 'Почему вы решили вступить в брак? ', reply_markup=markup)
        bot.register_next_step_handler(message, Expectations_of_marriage)


def Expectations_of_marriage1(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['Z' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    markup = types.ReplyKeyboardRemove()
    bot.send_message(message.chat.id, 'Каковы ваши ожидания от брака?', reply_markup=markup)
    bot.register_next_step_handler(message, Partner_support)


def Partner_support(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AA' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    markup = types.ReplyKeyboardMarkup()
    btn1 = types.KeyboardButton('Никогда')
    btn2 = types.KeyboardButton('Редко')
    btn3 = types.KeyboardButton('Иногда')
    btn4 = types.KeyboardButton('Часто')
    btn5 = types.KeyboardButton('Постоянно')
    markup.add(btn1, btn2, btn3, btn4, btn5)
    bot.send_message(message.chat.id, 'Чувствуете ли вы поддержку от вашего партнера в решении вступить в брак? ',
                     reply_markup=markup)
    bot.register_next_step_handler(message, Readiness_for_responsibility)


def Readiness_for_responsibility(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AB' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'никогда' or message.text.lower() == 'редко' or message.text.lower() == 'иногда' or message.text.lower() == 'часто' or message.text.lower() == 'постоянно':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Совсем не готов')
        btn2 = types.KeyboardButton('Не готов')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Готов')
        btn5 = types.KeyboardButton('Очень готов')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько вы готовы взять на себя ответственность за семейную жизнь? ',
                         reply_markup=markup)
        bot.register_next_step_handler(message, Joint_future)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Чувствуете ли вы поддержку от вашего партнера в решении вступить в брак? ',
                         reply_markup=markup)
        bot.register_next_step_handler(message, Readiness_for_responsibility)


def Joint_future(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AC' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'совсем не готов' or message.text.lower() == 'не готов' or message.text.lower() == 'нейтрально' or message.text.lower() == 'готов' or message.text.lower() == 'очень готов':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Обсуждали ли вы с партнером планы на совместное будущее?',
                         reply_markup=markup)
        bot.register_next_step_handler(message, Financial_stability)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Совсем не готов')
        btn2 = types.KeyboardButton('Не готов')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Готов')
        btn5 = types.KeyboardButton('Очень готов')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько вы готовы взять на себя ответственность за семейную жизнь? ',
                         reply_markup=markup)
        bot.register_next_step_handler(message, Joint_future)


def Financial_stability(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AD' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'никогда' or message.text.lower() == 'редко' or message.text.lower() == 'иногда' or message.text.lower() == 'часто' or message.text.lower() == 'постоянно':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень нестабильное')
        btn2 = types.KeyboardButton('Нестабильное')
        btn3 = types.KeyboardButton('Среднее')
        btn4 = types.KeyboardButton('Стабильное')
        btn5 = types.KeyboardButton('Очень стабильное')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как вы оцениваете ваше текущее финансовое состояние? ', reply_markup=markup)
        bot.register_next_step_handler(message, Joint_budget)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Обсуждали ли вы с партнером планы на совместное будущее?',
                         reply_markup=markup)
        bot.register_next_step_handler(message, Financial_stability)


def Joint_budget(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AE' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'очень нестабильное' or message.text.lower() == 'нестабильное' or message.text.lower() == 'среднее' or message.text.lower() == 'стабильное' or message.text.lower() == 'очень стабильное':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Да')
        btn2 = types.KeyboardButton('Нет')
        markup.add(btn1, btn2)
        bot.send_message(message.chat.id, 'Ведете ли вы совместный бюджет с партнером?', reply_markup=markup)
        bot.register_next_step_handler(message, Financial_conflicts)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Очень нестабильное')
        btn2 = types.KeyboardButton('Нестабильное')
        btn3 = types.KeyboardButton('Среднее')
        btn4 = types.KeyboardButton('Стабильное')
        btn5 = types.KeyboardButton('Очень стабильное')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Как вы оцениваете ваше текущее финансовое состояние? ', reply_markup=markup)
        bot.register_next_step_handler(message, Joint_budget)


def Financial_conflicts(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AF' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'да' or message.text.lower() == 'нет':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Возникают ли у вас конфликты из-за финансов?', reply_markup=markup)
        bot.register_next_step_handler(message, Personal_time)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Да')
        btn2 = types.KeyboardButton('Нет')
        markup.add(btn1, btn2)
        bot.send_message(message.chat.id, 'Ведете ли вы совместный бюджет с партнером?', reply_markup=markup)
        bot.register_next_step_handler(message, Financial_conflicts)


def Personal_time(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AG' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'никогда' or message.text.lower() == 'редко' or message.text.lower() == 'иногда' or message.text.lower() == 'часто' or message.text.lower() == 'постоянно':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Меньше 1 часа в день')
        btn2 = types.KeyboardButton('1-2 часа в день')
        btn3 = types.KeyboardButton('3-4 часа в день')
        btn4 = types.KeyboardButton('более 4 часов в день')
        markup.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, 'Сколько времени вы уделяете личным увлечениям и хобби?', reply_markup=markup)
        bot.register_next_step_handler(message, Skills_development)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Никогда')
        btn2 = types.KeyboardButton('Редко')
        btn3 = types.KeyboardButton('Иногда')
        btn4 = types.KeyboardButton('Часто')
        btn5 = types.KeyboardButton('Постоянно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Возникают ли у вас конфликты из-за финансов?', reply_markup=markup)
        bot.register_next_step_handler(message, Personal_time)


def Skills_development(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AH' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'меньше 1 часа в день' or message.text.lower() == '1-2 часа в день' or message.text.lower() == '3-4 часа в день' or message.text.lower() == 'более 4 часов в день':
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Совсем не важно')
        btn2 = types.KeyboardButton('Не важно')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Важно')
        btn5 = types.KeyboardButton('Очень важно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько важно для вас развитие личных и профессиональных навыков',
                         reply_markup=markup)
        bot.register_next_step_handler(message, Comments)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Меньше 1 часа в день')
        btn2 = types.KeyboardButton('1-2 часа в день')
        btn3 = types.KeyboardButton('3-4 часа в день')
        btn4 = types.KeyboardButton('более 4 часов в день')
        markup.add(btn1, btn2, btn3, btn4)
        bot.send_message(message.chat.id, 'Сколько времени вы уделяете личным увлечениям и хобби?', reply_markup=markup)
        bot.register_next_step_handler(message, Skills_development)


def Comments(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AI' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    if message.text.lower() == 'совсем не важно' or message.text.lower() == 'не важно' or message.text.lower() == 'нейтрально' or message.text.lower() == 'важно' or message.text.lower() == 'очень важно':
        markup = types.ReplyKeyboardRemove()
        bot.send_message(message.chat.id, 'Каковы ваши ожидания от вашего взаимодействия с психологом-консультантом?',
                         reply_markup=markup)
        bot.register_next_step_handler(message, end)
    else:
        markup = types.ReplyKeyboardMarkup()
        btn1 = types.KeyboardButton('Совсем не важно')
        btn2 = types.KeyboardButton('Не важно')
        btn3 = types.KeyboardButton('Нейтрально')
        btn4 = types.KeyboardButton('Важно')
        btn5 = types.KeyboardButton('Очень важно')
        markup.add(btn1, btn2, btn3, btn4, btn5)
        bot.send_message(message.chat.id, 'Насколько важно для вас развитие личных и профессиональных навыков',
                         reply_markup=markup)
        bot.register_next_step_handler(message, Comments)



def end(message):
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AJ' + ws['A1'].value] = message.text
    wb.save(fn)
    wb.close()
    #
    # использование этого номера
    fn = 'tablich.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    ws['AK' + ws['A1'].value] = '@' + message.from_user.username
    wb.save(fn)
    wb.close()
    #
    markup = types.ReplyKeyboardRemove()
    bot.send_message(message.from_user.id, f'Спасибо, что присоединились к нашему чат-боту в Telegram! 👩',
    f'‍⚕️Наши психологи готовы помочь вам разобраться в ваших брачно-семейных отношениях и найти пути к их улучшению.',
    f'Не стесняйтесь задавать вопросы и делиться своими мыслями — мы здесь, чтобы поддержать вас🫂Начните свой путь к более здоровым и счастливым отношениям уже сегодня!',
    f'Перейдите по ссылке @Fam_Helper, добавляйтесь в чат, чтобы начать общение с нашими специалистами. Вместе мы сможем добиться больших изменений📩', reply_markup=markup)


bot.polling(non_stop=True, interval=0)